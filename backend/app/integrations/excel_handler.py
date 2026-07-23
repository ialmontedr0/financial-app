"""Excel import/export handler for transactions."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
import pandas as pd

EXPECTED_COLUMNS_XLSX = {
    "date": ["date", "fecha", "effective_date", "transaction_date"],
    "description": ["description", "descripcion", "desc", "memo", "concepto"],
    "amount": ["amount", "monto", "importe", "valor"],
    "type": ["type", "tipo", "transaction_type"],
    "category": ["category", "categoria"],
    "account": ["account", "cuenta"],
    "currency": ["currency", "moneda"],
    "notes": ["notes", "notas"],
}

TYPE_ALIASES = {
    "ingreso": "income",
    "gasto": "expense",
    "expense": "expense",
    "income": "income",
    "transfer": "transfer",
    "transferencia": "transfer",
    "debit": "expense",
    "credit": "income",
    "cargo": "expense",
    "abono": "income",
}


def normalize_column(col: str) -> str:
    col_lower = col.strip().lower()
    for std, aliases in EXPECTED_COLUMNS_XLSX.items():
        if col_lower in aliases:
            return std
    return col_lower


def parse_excel(file_content: bytes) -> pd.DataFrame:
    """Parse Excel (.xlsx) content into a DataFrame."""
    df = pd.read_excel(
        io.BytesIO(file_content),
        engine="openpyxl",
        dtype=str,
        keep_default_na=False,
    )

    df.columns = [normalize_column(c) for c in df.columns]
    return df


def validate_rows_xlsx(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Validate Excel rows (same logic as CSV)."""
    from app.integrations.csv_handler import validate_rows

    return validate_rows(df)


def df_to_transactions_xlsx(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Convert Excel DataFrame to transaction dicts."""
    from app.integrations.csv_handler import df_to_transactions

    return df_to_transactions(df)


def transactions_to_excel(transactions: list[dict[str, Any]]) -> bytes:
    """Convert transactions list to Excel bytes (.xlsx)."""
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Description", "Amount", "Type", "Category", "Account", "Currency", "Notes"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for tx in transactions:
        ws.append([
            tx.get("date", ""),
            tx.get("description", ""),
            tx.get("amount", 0),
            tx.get("type", ""),
            tx.get("category", ""),
            tx.get("account", ""),
            tx.get("currency", "DOP"),
            tx.get("notes", ""),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
